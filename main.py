import openpyxl

# Excelファイルの読み込み
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.active

# データの読み込み
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(list(row))  # 各行をリストに変換して追加する

# データの加工や分析
# 例: データを加工して新しい列を追加する
for row in data:
    row.append(row[1] * 2)  # 2番目の列の値を2倍して新しい列に追加する

# 加工したデータを書き込む
for i, row in enumerate(data, start=1):
    for j, value in enumerate(row, start=1):
        sheet.cell(row=i, column=j, value=value)

# Excelファイルの保存
workbook.save('output.xlsx')
